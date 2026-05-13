#!/usr/bin/env python3
"""
Normalize Sve_Opstine_Spojeno 8.maj.2026.xlsx into a flat, canonical
one-row-per-polling-station dataset.

Reads the block-structured workbook (OPŠTINA: <name> headers + per-opstina
table) and emits:
  - data/sve_opstine_normalized.csv
  - data/sve_opstine_normalized.json
  - data/sve_opstine_normalized.xlsx
  - data/sve_opstine_normalize_report.md

The coordinate column has 8+ formats (decimal, DMS, comma-decimal+N/E,
embedded URLs, etc.). All are parsed into a single (lat, lon) pair with a
`coord_source_format` tag. Out-of-bbox values are nulled out and reported.

Pure offline pass — see scripts/resolve_goo_gl.py for the optional online
follow-up that recovers coords from maps.app.goo.gl shortlinks.

Usage:
    python3 scripts/normalize_sve_opstine.py
    python3 scripts/normalize_sve_opstine.py --xlsx X --out-dir D
"""

import argparse
import csv
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl import Workbook


HERE = Path(__file__).resolve().parent
REPO = HERE.parent
DEFAULT_XLSX = REPO / "Sve_Opstine_Spojeno 8.maj.2026.xlsx"
DEFAULT_OUT_DIR = REPO / "data"
LOCALITIES_JSON = REPO / "polling_stations_86.json"

SERBIA_BBOX_LAT = (41.5, 46.5)
SERBIA_BBOX_LON = (18.5, 23.5)

# Cyrillic <-> Latin (Serbian, Vuk/Gaj). Digraphs first.
CYR_TO_LAT_DIGRAPH = {"Љ": "Lj", "Њ": "Nj", "Џ": "Dž"}
CYR_TO_LAT_MAP = {
    "А": "A", "Б": "B", "В": "V", "Г": "G", "Д": "D", "Ђ": "Đ", "Е": "E",
    "Ж": "Ž", "З": "Z", "И": "I", "Ј": "J", "К": "K", "Л": "L", "М": "M",
    "Н": "N", "О": "O", "П": "P", "Р": "R", "С": "S", "Т": "T", "Ћ": "Ć",
    "У": "U", "Ф": "F", "Х": "H", "Ц": "C", "Ч": "Č", "Ш": "Š",
}
LAT_TO_CYR_DIGRAPH = {"LJ": "Љ", "NJ": "Њ", "DŽ": "Џ", "DJ": "Ђ"}
LAT_TO_CYR_MAP = {
    "A": "А", "B": "Б", "V": "В", "G": "Г", "D": "Д", "Đ": "Ђ", "E": "Е",
    "Ž": "Ж", "Z": "З", "I": "И", "J": "Ј", "K": "К", "L": "Л", "M": "М",
    "N": "Н", "O": "О", "P": "П", "R": "Р", "S": "С", "T": "Т", "Ć": "Ћ",
    "U": "У", "F": "Ф", "H": "Х", "C": "Ц", "Č": "Ч", "Š": "Ш",
}


def cyr_to_lat(s: str) -> str:
    if not s:
        return s
    out = []
    i = 0
    while i < len(s):
        ch = s[i]
        up = ch.upper()
        if up in CYR_TO_LAT_DIGRAPH:
            mapped = CYR_TO_LAT_DIGRAPH[up]
            out.append(mapped if ch.isupper() else mapped.lower())
        elif up in CYR_TO_LAT_MAP:
            mapped = CYR_TO_LAT_MAP[up]
            out.append(mapped if ch.isupper() else mapped.lower())
        else:
            out.append(ch)
        i += 1
    return "".join(out)


def lat_to_cyr(s: str) -> str:
    if not s:
        return s
    out = []
    i = 0
    while i < len(s):
        # Check digraphs (case-insensitive)
        two = s[i:i + 2].upper()
        if two in LAT_TO_CYR_DIGRAPH:
            mapped = LAT_TO_CYR_DIGRAPH[two]
            out.append(mapped if s[i].isupper() else mapped.lower())
            i += 2
            continue
        ch = s[i]
        up = ch.upper()
        if up in LAT_TO_CYR_MAP:
            mapped = LAT_TO_CYR_MAP[up]
            out.append(mapped if ch.isupper() else mapped.lower())
        else:
            out.append(ch)
        i += 1
    return "".join(out)


def is_cyrillic(s: str) -> bool:
    return any("А" <= c <= "џ" or c in "ЂЉЊЋЏ" for c in s)


# Hand-curated overrides for opstinas where naïve transliteration won't
# match polling_stations_86.json's locality names (Beograd/Niš sub-opstinas,
# "Град X" / "X - ГРАД" patterns, special compound names).
# Maps the *raw* OPŠTINA: header (verbatim) -> JSON locality name.
OPSTINA_OVERRIDES = {
    # Градска општина <X> (Beograd)
    "Градска општина Барајево": "БЕОГРАД-БАРАЈЕВО",
    "Градска општина Вождовац": "БЕОГРАД-ВОЖДОВАЦ",
    "Градска општина Врачар": "БЕОГРАД-ВРАЧАР",
    "Градска општина Гроцка": "БЕОГРАД-ГРОЦКА",
    "Градска општина Звездара": "БЕОГРАД-ЗВЕЗДАРА",
    "Градска општина Земун": "БЕОГРАД-ЗЕМУН",
    "Градска општина Лазаревац": "БЕОГРАД-ЛАЗАРЕВАЦ",
    "Градска општина Младеновац": "БЕОГРАД-МЛАДЕНОВАЦ",
    "Градска општина Нови Београд": "БЕОГРАД-НОВИ БЕОГРАД",
    "Градска општина Обреновац": "БЕОГРАД-ОБРЕНОВАЦ",
    "Градска општина Палилула (Београд)": "БЕОГРАД-ПАЛИЛУЛА",
    "Градска општина Раковица": "БЕОГРАД-РАКОВИЦА",
    "Градска општина Савски Венац": "БЕОГРАД-САВСКИ ВЕНАЦ",
    "Градска општина Сопот": "БЕОГРАД-СОПОТ",
    "Градска општина Стари Град": "БЕОГРАД-СТАРИ ГРАД",
    "Градска општина Сурчин": "БЕОГРАД-СУРЧИН",
    "Градска општина Чукарица": "БЕОГРАД-ЧУКАРИЦА",
    # Градска општина <X> (Niš)
    "Градска општина Медијана": "НИШ-МЕДИЈАНА",
    "Градска општина Нишка Бања": "НИШ-НИШКА БАЊА",
    "Градска општина Палилула (Ниш)": "НИШ-ПАЛИЛУЛА",
    "Градска општина Пантелеј": "НИШ-ПАНТЕЛЕЈ",
    "Градска општина Црвени крст": "НИШ-ЦРВЕНИ КРСТ",
    # Other "Градска општина" entries that map to compound names
    "Градска општина Костолац": "ПОЖАРЕВАЦ-КОСТОЛАЦ",
    "Градска општина Севојно": "УЖИЦЕ-СЕВОЈНО",
    "Градска општина Врањска Бања": "ВРАЊЕ-ВРАЊСКА БАЊА",
    # Град <X> -> X - ГРАД
    "Град Јагодина": "ЈАГОДИНА - ГРАД",
    "Град Бор": "БОР - ГРАД",
    "Град Ваљево": "ВАЉЕВО - ГРАД",
    "Град Врање": "ВРАЊЕ - ГРАД",
    "Град Зајечар": "ЗАЈЕЧАР - ГРАД",
    "Град Зрењанин": "ЗРЕЊАНИН - ГРАД",
    "Град Кикинда": "КИКИНДА - ГРАД",
    "Град Крагујевац": "КРАГУЈЕВАЦ - ГРАД",
    "Град Краљево": "КРАЉЕВО - ГРАД",
    "Град Крушевац": "КРУШЕВАЦ - ГРАД",
    "Град Лесковац": "ЛЕСКОВАЦ - ГРАД",
    "Град Лозница": "ЛОЗНИЦА - ГРАД",
    "Град Нови Пазар": "НОВИ ПАЗАР - ГРАД",
    "Град Нови Сад": "НОВИ САД - ГРАД",
    "Град Пирот": "ПИРОТ - ГРАД",
    "Град Пожаревац": "ПОЖАРЕВАЦ",
    "Град Смедерево": "СМЕДЕРЕВО - ГРАД",
    "Град Сомбор": "СОМБОР - ГРАД",
    "Град Сремска Митровица": "СРЕМСКА МИТРОВИЦА - ГРАД",
    "Град Суботица": "СУБОТИЦА - ГРАД",
    "Град Ужице": "УЖИЦЕ - ГРАД",
    "Град Чачак": "ЧАЧАК - ГРАД",
    "Град Шабац": "ШАБАЦ - ГРАД",
    # Pančevo / Vršac / Niš proper don't appear in xlsx as opstina blocks;
    # but Pančevo does:
    "Pančevo": "ПАНЧЕВО - ГРАД",
    "Vršac": "ВРШАЦ - ГРАД",
    # Plain Cyrillic name that the JSON suffixes with " - ГРАД"
    "Прокупље": "ПРОКУПЉЕ - ГРАД",
}


def canonical_opstina(raw: str, json_names: set) -> tuple:
    """Return (opstina_cyr, opstina_lat) for the raw OPŠTINA header string.

    Strategy:
      1. Hand-curated override (handles Beograd/Niš and city-with-suffix).
      2. If already in JSON set (exact or uppercase) -> use it.
      3. Transliterate to Cyrillic uppercase; if it matches JSON -> use it.
      4. Otherwise return (uppercase-cyrillic-best-effort, titlecase-latin).
         Caller logs unmatched.
    """
    raw = raw.strip()
    if raw in OPSTINA_OVERRIDES:
        cyr = OPSTINA_OVERRIDES[raw]
        lat = cyr_to_lat(cyr)
        return cyr, lat

    # If raw is Latin, transliterate to Cyrillic
    if is_cyrillic(raw):
        cyr = raw.upper()
        lat = cyr_to_lat(cyr).title()
    else:
        cyr = lat_to_cyr(raw).upper()
        lat = raw

    # Try exact JSON match
    if cyr in json_names:
        return cyr, cyr_to_lat(cyr).title()
    upper_match = next((n for n in json_names if n == cyr), None)
    if upper_match:
        return upper_match, cyr_to_lat(upper_match).title()
    return cyr, lat


# ---------------------------------------------------------------------------
# Coordinate parser
# ---------------------------------------------------------------------------

_URL_AT_RE = re.compile(r"@(-?\d+\.\d+),(-?\d+\.\d+)")
_DECIMAL_DEG_DIR_RE = re.compile(
    r"(-?\d+(?:[.,]\d+)?)\s*°\s*([NSEW])", re.IGNORECASE
)
_DMS_RE = re.compile(
    r"(-?\d+)\s*°\s*(\d+)\s*['′]\s*(\d+(?:\.\d+)?)\s*[\"″]?\s*([NSEW]?)",
    re.IGNORECASE,
)
_NS_DECIMAL_RE = re.compile(
    r"(-?\d+(?:[.,]\d+)?)\s*([NSEW])", re.IGNORECASE
)
_PLAIN_DECIMAL_RE = re.compile(r"-?\d+\.\d+")


def _swap_if_lon_first(lat: float, lon: float, first_dir: str) -> tuple:
    if first_dir.upper() in ("E", "W"):
        return lon, lat
    return lat, lon


def _to_decimal(deg: str, mins: str, secs: str, direction: str) -> tuple:
    val = abs(float(deg)) + float(mins) / 60.0 + float(secs) / 3600.0
    sign = -1 if (deg.lstrip().startswith("-") or direction.upper() in ("S", "W")) else 1
    return val * sign, direction


def parse_coord(raw):
    """Parse a coordinate cell into (lat, lon, format_tag).

    Format tags:
      empty, url_embedded, url_no_coord, shortlink_pending,
      decimal_simple, dms, decimal_dir, comma_dec_comma_sep,
      dms_decimal_deg, dms_unparsed, unparsed
    """
    if raw is None:
        return None, None, "empty"
    if isinstance(raw, (int, float)):
        return None, None, "unparsed"
    s = str(raw).strip()
    if not s:
        return None, None, "empty"

    # URLs
    if re.search(r"https?://", s, re.IGNORECASE):
        m = _URL_AT_RE.search(s)
        if m:
            return float(m.group(1)), float(m.group(2)), "url_embedded"
        if "maps.app.goo.gl" in s.lower():
            return None, None, "shortlink_pending"
        return None, None, "url_no_coord"

    # Collapse internal whitespace + newlines
    s_clean = re.sub(r"\s+", " ", s.replace("\r", " ").replace("\n", " ")).strip()

    if "°" in s_clean:
        # Pattern A: decimal-degrees with direction (e.g. "45,39799° N, ...")
        a_matches = _DECIMAL_DEG_DIR_RE.findall(s_clean)
        if len(a_matches) >= 2:
            v1 = float(a_matches[0][0].replace(",", "."))
            d1 = a_matches[0][1].upper()
            v2 = float(a_matches[1][0].replace(",", "."))
            d2 = a_matches[1][1].upper()
            sign1 = -1 if d1 in ("S", "W") else 1
            sign2 = -1 if d2 in ("S", "W") else 1
            lat, lon = _swap_if_lon_first(v1 * sign1, v2 * sign2, d1)
            return lat, lon, "dms_decimal_deg"

        # Pattern B: DMS
        b_matches = _DMS_RE.findall(s_clean)
        if len(b_matches) >= 2:
            v0 = _to_decimal(*b_matches[0])
            v1 = _to_decimal(*b_matches[1])
            lat, lon = _swap_if_lon_first(v0[0], v1[0], v0[1])
            return lat, lon, "dms"

        return None, None, "dms_unparsed"

    # NS/EW without °
    ns_matches = _NS_DECIMAL_RE.findall(s_clean)
    if len(ns_matches) >= 2:
        v1 = float(ns_matches[0][0].replace(",", "."))
        d1 = ns_matches[0][1].upper()
        v2 = float(ns_matches[1][0].replace(",", "."))
        d2 = ns_matches[1][1].upper()
        sign1 = -1 if d1 in ("S", "W") else 1
        sign2 = -1 if d2 in ("S", "W") else 1
        lat, lon = _swap_if_lon_first(v1 * sign1, v2 * sign2, d1)
        return lat, lon, "decimal_dir"

    # Comma decimal + comma separator (4 commas, no dots): "44,4175396, 20,4457139"
    if s_clean.count(",") == 3 and "." not in s_clean:
        parts = [p.strip() for p in s_clean.split(",")]
        if all(p.lstrip("-").isdigit() for p in parts):
            try:
                lat = float(parts[0] + "." + parts[1])
                lon = float(parts[2] + "." + parts[3])
                return lat, lon, "comma_dec_comma_sep"
            except ValueError:
                pass

    # Plain decimal pair
    nums = _PLAIN_DECIMAL_RE.findall(s_clean)
    if len(nums) >= 2:
        return float(nums[0]), float(nums[1]), "decimal_simple"

    return None, None, "unparsed"


# ---------------------------------------------------------------------------
# Aux field normalization
# ---------------------------------------------------------------------------

SIGNAL_MAP = {
    "Раде позиви, поруке и интернет": "ok",
    "Раде само позиви и/или СМС поруке": "voice_sms_only",
    "Нема сигнала": "none",
}


def normalize_confirm(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    low = s.lower()
    if low in ("true", "1. da", "1.da", "da", "1"):
        return True
    if low in ("false", "2. ne", "2.ne", "ne", "0"):
        return False
    return None  # unknown


def normalize_signal(v):
    if v is None:
        return None, None
    s = str(v).strip()
    if not s:
        return None, None
    return SIGNAL_MAP.get(s), s


# ---------------------------------------------------------------------------
# Row detection
# ---------------------------------------------------------------------------

_RB_INT_STR_RE = re.compile(r"^\s*(\d+)\s*\.*\s*$")
_RB_COMMA_RE = re.compile(r"^\s*(\d+)\s*,\s*(\d+)\s*$")
_RB_BM_RE = re.compile(r"^\s*BM\s*0*(\d+)\s*$", re.IGNORECASE)


def parse_rb(v):
    if v is None:
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        if v.is_integer():
            return int(v)
        return None
    if isinstance(v, str):
        m = _RB_INT_STR_RE.match(v)
        if m:
            return int(m.group(1))
        m = _RB_BM_RE.match(v)
        if m:
            return int(m.group(1))
        m = _RB_COMMA_RE.match(v)
        if m:
            try:
                f = float(f"{m.group(1)}.{m.group(2)}")
                if f.is_integer():
                    return int(f)
            except ValueError:
                pass
    return None


def normalize_name(raw):
    """Return (name_cyr, name_lat). If cell is Cyrillic\\nLatin (Sjenica
    style) where line2 is the transliteration of line1, split them.
    Otherwise return cleaned single string in name_cyr and None in name_lat.
    """
    if raw is None:
        return None, None
    s = str(raw)
    if "\n" in s:
        lines = [ln.strip() for ln in s.split("\n") if ln.strip()]
        if len(lines) == 2 and is_cyrillic(lines[0]) and not is_cyrillic(lines[1]):
            # Heuristic: confirm line2 looks like transliteration of line1
            expected = cyr_to_lat(lines[0]).lower().replace(" ", "")
            actual = lines[1].lower().replace(" ", "")
            # Drop diacritics for a lenient match
            def fold(x):
                return "".join(
                    c for c in unicodedata.normalize("NFKD", x)
                    if not unicodedata.combining(c)
                )
            if fold(expected) == fold(actual) or len(expected) == len(actual):
                return lines[0], lines[1]
    cleaned = re.sub(r"\s+", " ", s.replace("\r", " ").replace("\n", " ")).strip()
    return cleaned or None, None


def normalize_text(raw):
    if raw is None:
        return None
    s = str(raw)
    cleaned = re.sub(r"[ \t]+", " ", s.replace("\r", " ")).strip()
    cleaned = re.sub(r"\n\s*", " ", cleaned)
    return cleaned or None


def extract_gmaps_url_coord(url):
    if not url or not isinstance(url, str):
        return None, None
    m = _URL_AT_RE.search(url)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None, None


# ---------------------------------------------------------------------------
# Walk the workbook
# ---------------------------------------------------------------------------

_HEADER_FIRST_CELL = {"РБ", "РЕД. БР.", "РЕД.БР.", "РЕДНИ БРОЈ", "RED. BR.", "RED.BR."}


def _is_header_row(row):
    """A per-opstina header row carries the column titles (РБ / НАЗИВ / ...).
    Detect by looking at first cell or by spotting 'НАЗИВ' anywhere in row.
    """
    first = row[0]
    if isinstance(first, str) and first.strip().upper() in _HEADER_FIRST_CELL:
        return True
    # Fallback: if col 1 says НАЗИВ ГЛАСАЧКОГ ... it's a header
    second = row[1]
    if isinstance(second, str) and "НАЗИВ" in second.upper():
        return True
    return False


def walk_rows(ws):
    """Yield (excel_row, opstina_raw, row_tuple) for every data row."""
    current_opst = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        first = row[0]
        if isinstance(first, str) and first.strip().upper().startswith("OPŠTINA"):
            current_opst = first.split(":", 1)[1].strip()
            continue
        if _is_header_row(row):
            continue
        if current_opst is None:
            continue
        # Skip fully-empty rows
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        # Skip stray non-data rows (e.g. blank in col 0 and col 1)
        if (row[0] is None and (row[1] is None or (isinstance(row[1], str) and not row[1].strip()))):
            continue
        yield i, current_opst, row


def emit_record(excel_row, opstina_raw, row, json_names):
    opst_cyr, opst_lat = canonical_opstina(opstina_raw, json_names)
    rb = parse_rb(row[0])
    name_cyr, name_lat = normalize_name(row[1])
    address = normalize_text(row[2])
    area = normalize_text(row[3])
    coord_raw = None
    if row[4] is not None:
        coord_raw = str(row[4])
    lat, lon, fmt = parse_coord(row[4])

    # Sanity bbox check
    out_of_bbox = False
    if lat is not None and lon is not None:
        if not (SERBIA_BBOX_LAT[0] <= lat <= SERBIA_BBOX_LAT[1]) or \
           not (SERBIA_BBOX_LON[0] <= lon <= SERBIA_BBOX_LON[1]):
            out_of_bbox = True
            lat, lon = None, None
            fmt = "out_of_bbox"

    gmaps_url = row[5] if isinstance(row[5], str) and row[5].strip() else None
    if gmaps_url:
        gmaps_url = gmaps_url.strip()

    # Fallback: gmaps URL has @lat,lon and we don't have coords yet
    if lat is None and gmaps_url:
        gl_lat, gl_lon = extract_gmaps_url_coord(gmaps_url)
        if gl_lat is not None:
            if SERBIA_BBOX_LAT[0] <= gl_lat <= SERBIA_BBOX_LAT[1] and \
               SERBIA_BBOX_LON[0] <= gl_lon <= SERBIA_BBOX_LON[1]:
                lat, lon = gl_lat, gl_lon
                fmt = "from_gmaps_url"

    map_confirmed = normalize_confirm(row[6])
    note = normalize_text(row[7])
    sig_mts, sig_mts_raw = normalize_signal(row[8])
    sig_yt, sig_yt_raw = normalize_signal(row[9])
    sig_a1, sig_a1_raw = normalize_signal(row[10])
    extra_note = normalize_text(row[11]) if len(row) > 11 else None

    return {
        "opstina_cyr": opst_cyr,
        "opstina_lat": opst_lat,
        "opstina_raw": opstina_raw,
        "rb": rb,
        "name_cyr": name_cyr,
        "name_lat": name_lat,
        "address": address,
        "area": area,
        "lat": lat,
        "lon": lon,
        "coord_source_format": fmt,
        "coord_raw": coord_raw,
        "gmaps_url": gmaps_url,
        "gmaps_url_resolved": None,
        "map_confirmed": map_confirmed,
        "note": note,
        "signal_mts": sig_mts,
        "signal_mts_raw": sig_mts_raw if sig_mts is None and sig_mts_raw else None,
        "signal_yettel": sig_yt,
        "signal_yettel_raw": sig_yt_raw if sig_yt is None and sig_yt_raw else None,
        "signal_a1": sig_a1,
        "signal_a1_raw": sig_a1_raw if sig_a1 is None and sig_a1_raw else None,
        "extra_note": extra_note,
        "row_excel": excel_row,
        "_out_of_bbox": out_of_bbox,
    }


# ---------------------------------------------------------------------------
# Emitters
# ---------------------------------------------------------------------------

CSV_COLUMNS = [
    "row_excel", "opstina_cyr", "opstina_lat", "opstina_raw", "rb",
    "name_cyr", "name_lat", "address", "area",
    "lat", "lon", "coord_source_format", "coord_raw",
    "gmaps_url", "gmaps_url_resolved",
    "map_confirmed", "note",
    "signal_mts", "signal_yettel", "signal_a1",
    "signal_mts_raw", "signal_yettel_raw", "signal_a1_raw",
    "extra_note",
]


def write_csv(records, path):
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        w.writeheader()
        for r in records:
            w.writerow(r)


def write_json(records, path):
    public = []
    for r in records:
        public.append({k: r.get(k) for k in CSV_COLUMNS})
    with path.open("w", encoding="utf-8") as f:
        json.dump(public, f, ensure_ascii=False, indent=2)


def write_xlsx(records, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "polling_stations"
    ws.append(CSV_COLUMNS)
    for r in records:
        ws.append([r.get(c) for c in CSV_COLUMNS])
    wb.save(path)


def write_report(records, json_names, source_xlsx, out_path):
    by_opst = defaultdict(list)
    fmt_counts = Counter()
    for r in records:
        by_opst[r["opstina_cyr"]].append(r)
        fmt_counts[r["coord_source_format"]] += 1

    total = len(records)
    with_coords = sum(1 for r in records if r["lat"] is not None)

    unmatched_opst = sorted({
        r["opstina_cyr"] for r in records if r["opstina_cyr"] not in json_names
    })
    missing_rb = [r for r in records if r["rb"] is None]
    unparsed = [r for r in records if r["coord_source_format"] in
                ("unparsed", "dms_unparsed", "url_no_coord", "shortlink_pending")
                and r["coord_raw"] is not None]
    out_of_bbox = [
        r for r in records
        if r.get("_out_of_bbox") or r.get("coord_source_format") == "out_of_bbox"
    ]
    unknown_signal = [
        r for r in records
        if r.get("signal_mts_raw") or r.get("signal_yettel_raw") or r.get("signal_a1_raw")
    ]

    lines = []
    lines.append(f"# Sve_Opstine_Spojeno normalize report\n")
    lines.append(f"Source: `{source_xlsx.name}`")
    lines.append("")
    lines.append(f"- Total rows emitted: **{total}**")
    lines.append(f"- Rows with valid `lat`/`lon`: **{with_coords}** ({with_coords*100//max(total,1)}%)")
    lines.append(f"- Opstinas covered: **{len(by_opst)}**")
    lines.append(f"- Opstinas not found in `polling_stations_86.json`: **{len(unmatched_opst)}**")
    lines.append(f"- Rows missing РБ: **{len(missing_rb)}**")
    lines.append(f"- Rows with unparsed coord text: **{len(unparsed)}**")
    lines.append(f"- Rows with out-of-bbox coords (nulled out): **{len(out_of_bbox)}**")
    lines.append(f"- Rows with unknown signal value: **{len(unknown_signal)}**")
    lines.append("")
    lines.append("## Coord format histogram\n")
    for fmt, n in fmt_counts.most_common():
        lines.append(f"- `{fmt}`: {n}")
    lines.append("")

    lines.append("## Per-opstina coverage\n")
    lines.append("| Opstina (Cyr) | Rows | With coords | Matched JSON |")
    lines.append("|---|---:|---:|:---:|")
    for opst in sorted(by_opst):
        rs = by_opst[opst]
        n_coord = sum(1 for r in rs if r["lat"] is not None)
        match = "✓" if opst in json_names else "✗"
        lines.append(f"| {opst} | {len(rs)} | {n_coord} | {match} |")
    lines.append("")

    if unmatched_opst:
        lines.append("## Opstinas NOT matched in polling_stations_86.json\n")
        for o in unmatched_opst:
            lines.append(f"- `{o}`")
        lines.append("")

    if unparsed:
        lines.append("## Unparsed coord cells\n")
        for r in unparsed[:200]:
            lines.append(
                f"- row {r['row_excel']} ({r['opstina_cyr']}, РБ={r['rb']}): "
                f"`{r['coord_source_format']}` — `{(r['coord_raw'] or '')[:120]}`"
            )
        if len(unparsed) > 200:
            lines.append(f"\n_(+{len(unparsed) - 200} more)_")
        lines.append("")

    if missing_rb:
        lines.append("## Rows missing РБ\n")
        for r in missing_rb[:50]:
            lines.append(
                f"- row {r['row_excel']} ({r['opstina_cyr']}): name=`{(r['name_cyr'] or '')[:60]}`"
            )
        if len(missing_rb) > 50:
            lines.append(f"\n_(+{len(missing_rb) - 50} more)_")
        lines.append("")

    if out_of_bbox:
        lines.append("## Out-of-bbox coords (nulled)\n")
        for r in out_of_bbox:
            lines.append(
                f"- row {r['row_excel']} ({r['opstina_cyr']}, РБ={r['rb']}): "
                f"raw=`{r['coord_raw']}`"
            )
        lines.append("")

    if unknown_signal:
        lines.append("## Unknown signal-strength values\n")
        seen = set()
        for r in unknown_signal:
            for k in ("signal_mts_raw", "signal_yettel_raw", "signal_a1_raw"):
                v = r.get(k)
                if v and v not in seen:
                    seen.add(v)
                    lines.append(f"- {k}: `{v}` (e.g. row {r['row_excel']})")
        lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    out_dir = Path(args.out_dir)
    if not xlsx_path.exists():
        sys.exit(f"not found: {xlsx_path}")
    out_dir.mkdir(parents=True, exist_ok=True)

    # Load JSON locality names for cross-check
    json_names = set()
    if LOCALITIES_JSON.exists():
        with LOCALITIES_JSON.open(encoding="utf-8") as f:
            j = json.load(f)
        json_names = {loc["name"] for loc in j.get("localities", [])}

    print(f"Loading {xlsx_path.name}...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []
    for excel_row, opst_raw, row in walk_rows(ws):
        rec = emit_record(excel_row, opst_raw, row, json_names)
        records.append(rec)

    print(f"  {len(records)} rows emitted")
    n_coord = sum(1 for r in records if r["lat"] is not None)
    print(f"  {n_coord} have lat/lon")

    csv_out = out_dir / "sve_opstine_normalized.csv"
    json_out = out_dir / "sve_opstine_normalized.json"
    xlsx_out = out_dir / "sve_opstine_normalized.xlsx"
    report_out = out_dir / "sve_opstine_normalize_report.md"

    write_csv(records, csv_out)
    write_json(records, json_out)
    write_xlsx(records, xlsx_out)
    write_report(records, json_names, xlsx_path, report_out)

    print(f"  wrote {csv_out}")
    print(f"  wrote {json_out}")
    print(f"  wrote {xlsx_out}")
    print(f"  wrote {report_out}")


if __name__ == "__main__":
    main()
